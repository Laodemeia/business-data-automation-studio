from __future__ import annotations

import csv
import json
import re
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterable

from .translations import translate

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:  # pragma: no cover - handled with a user-facing error
    Workbook = None
    load_workbook = None


SUPPORTED_ENCODINGS = ("utf-8-sig", "utf-8", "cp1254", "latin-1")
SUPPORTED_DELIMITERS = (",", ";", "\t", "|")
SUPPORTED_FILE_TYPES = (".csv", ".txt", ".xlsx")


class DataStudioError(Exception):
    """Raised when a data file or workflow configuration is invalid."""


@dataclass
class ProcessingConfig:
    trim_whitespace: bool = True
    normalize_headers: bool = True
    remove_empty_rows: bool = True
    remove_duplicates: bool = True
    required_columns: list[str] = field(default_factory=list)
    filter_column: str = ""
    filter_mode: str = "contains"
    filter_value: str = ""
    column_renames: dict[str, str] = field(default_factory=dict)

    def to_dict(self) -> dict:
        return asdict(self)

    @classmethod
    def from_dict(cls, data: dict) -> "ProcessingConfig":
        allowed_keys = {item.name for item in cls.__dataclass_fields__.values()}
        cleaned = {key: value for key, value in data.items() if key in allowed_keys}
        return cls(**cleaned)


@dataclass(frozen=True)
class ProcessingIssue:
    source_row: int
    column: str
    message: str
    value: str = ""


@dataclass
class ProcessingResult:
    headers: list[str]
    rows: list[dict[str, str]]
    issues: list[ProcessingIssue]
    stats: dict[str, int]


def _make_unique(values: Iterable[str]) -> list[str]:
    counts: dict[str, int] = {}
    unique_values: list[str] = []

    for raw_value in values:
        value = raw_value or "column"
        counts[value] = counts.get(value, 0) + 1
        suffix = counts[value]
        unique_values.append(value if suffix == 1 else f"{value}_{suffix}")

    return unique_values


def normalize_header(value: str) -> str:
    cleaned = value.strip().lower()
    cleaned = re.sub(r"[^a-z0-9çğıöşü]+", "_", cleaned, flags=re.IGNORECASE)
    return cleaned.strip("_") or "column"


class DataProcessingEngine:
    def __init__(self, language: str = "en") -> None:
        self.language = language
        self.source_path: Path | None = None
        self.source_kind = ""
        self.encoding = "utf-8"
        self.delimiter = ","
        self.worksheet_name = ""
        self.worksheet_names: list[str] = []
        self.original_headers: list[str] = []
        self.original_rows: list[dict[str, str]] = []
        self.result: ProcessingResult | None = None

    def set_language(self, language: str) -> None:
        self.language = language

    def _t(self, key: str, **values) -> str:
        return translate(self.language, key, **values)

    def load_file(
        self,
        path: str | Path,
        *,
        worksheet: str | None = None,
    ) -> None:
        source = Path(path)
        if not source.is_file():
            raise DataStudioError(self._t("engine_file_missing"))

        suffix = source.suffix.lower()
        if suffix not in SUPPORTED_FILE_TYPES:
            supported = ", ".join(SUPPORTED_FILE_TYPES)
            raise DataStudioError(
                self._t("engine_unsupported_type", supported=supported)
            )

        if suffix == ".xlsx":
            self._load_excel(source, worksheet)
            return

        self._load_delimited_text(source)

    def _load_delimited_text(self, source: Path) -> None:
        text, encoding = self._read_text(source)
        delimiter = self._detect_delimiter(text)

        reader = csv.reader(text.splitlines(), delimiter=delimiter)
        raw_rows = list(reader)

        if not raw_rows:
            raise DataStudioError(self._t("engine_empty_file"))

        raw_headers = [header.strip() for header in raw_rows[0]]
        if not any(raw_headers):
            raise DataStudioError(self._t("engine_no_headers"))

        headers = _make_unique(raw_headers)
        rows: list[dict[str, str]] = []

        for raw_row in raw_rows[1:]:
            padded = raw_row + [""] * max(0, len(headers) - len(raw_row))
            rows.append(
                {
                    header: padded[index] if index < len(padded) else ""
                    for index, header in enumerate(headers)
                }
            )

        self.source_path = source
        self.source_kind = source.suffix.lower().lstrip(".")
        self.encoding = encoding
        self.delimiter = delimiter
        self.worksheet_name = ""
        self.worksheet_names = []
        self.original_headers = headers
        self.original_rows = rows
        self.result = None

    def _load_excel(self, source: Path, worksheet: str | None) -> None:
        if load_workbook is None:
            raise DataStudioError(self._t("engine_excel_dependency"))

        try:
            workbook = load_workbook(
                source,
                read_only=True,
                data_only=True,
            )
        except (OSError, ValueError, KeyError, InvalidFileException) as error:
            raise DataStudioError(
                self._t("engine_excel_read_error", error=error)
            )

        try:
            sheet_names = list(workbook.sheetnames)
            if not sheet_names:
                raise DataStudioError(self._t("engine_no_worksheets"))

            selected_name = worksheet or workbook.active.title
            if selected_name not in sheet_names:
                raise DataStudioError(
                    self._t(
                        "engine_worksheet_not_found",
                        sheet=selected_name,
                        available=", ".join(sheet_names),
                    )
                )

            sheet = workbook[selected_name]
            raw_rows = [
                [self._cell_to_text(value) for value in row]
                for row in sheet.iter_rows(values_only=True)
            ]
        finally:
            workbook.close()

        if not raw_rows:
            raise DataStudioError(self._t("engine_empty_worksheet"))

        raw_headers = [header.strip() for header in raw_rows[0]]
        if not any(raw_headers):
            raise DataStudioError(self._t("engine_no_worksheet_headers"))

        headers = _make_unique(raw_headers)
        rows: list[dict[str, str]] = []
        for raw_row in raw_rows[1:]:
            padded = raw_row + [""] * max(0, len(headers) - len(raw_row))
            rows.append(
                {
                    header: padded[index] if index < len(padded) else ""
                    for index, header in enumerate(headers)
                }
            )

        self.source_path = source
        self.source_kind = "xlsx"
        self.encoding = ""
        self.delimiter = ""
        self.worksheet_name = selected_name
        self.worksheet_names = sheet_names
        self.original_headers = headers
        self.original_rows = rows
        self.result = None

    def process(self, config: ProcessingConfig) -> ProcessingResult:
        if self.source_path is None:
            raise DataStudioError(self._t("engine_load_before_process"))

        headers, header_map = self._build_headers(config)
        required_columns = [
            header_map.get(column, column) for column in config.required_columns
        ]
        missing_columns = [
            column for column in required_columns if column not in headers
        ]
        if missing_columns:
            joined = ", ".join(missing_columns)
            raise DataStudioError(
                self._t("engine_required_not_found", columns=joined)
            )

        filter_column = header_map.get(config.filter_column, config.filter_column)
        if filter_column and filter_column not in headers:
            raise DataStudioError(
                self._t("engine_filter_not_found", column=filter_column)
            )

        processed_rows: list[dict[str, str]] = []
        issues: list[ProcessingIssue] = []
        empty_removed = 0
        filtered_removed = 0

        for source_index, source_row in enumerate(self.original_rows, start=2):
            row = {
                header_map[old_header]: self._clean_value(
                    source_row.get(old_header, ""),
                    config,
                )
                for old_header in self.original_headers
            }

            if config.remove_empty_rows and not any(row.values()):
                empty_removed += 1
                continue

            if filter_column and not self._matches_filter(
                row.get(filter_column, ""),
                config.filter_mode,
                config.filter_value,
            ):
                filtered_removed += 1
                continue

            for required_column in required_columns:
                if not row.get(required_column, "").strip():
                    issues.append(
                        ProcessingIssue(
                            source_row=source_index,
                            column=required_column,
                            message=self._t("engine_missing_required"),
                        )
                    )

            processed_rows.append(row)

        duplicate_removed = 0
        if config.remove_duplicates:
            processed_rows, duplicate_removed = self._remove_duplicates(
                headers,
                processed_rows,
            )

        stats = {
            "source_rows": len(self.original_rows),
            "output_rows": len(processed_rows),
            "empty_rows_removed": empty_removed,
            "filtered_rows_removed": filtered_removed,
            "duplicates_removed": duplicate_removed,
            "validation_issues": len(issues),
        }

        self.result = ProcessingResult(
            headers=headers,
            rows=processed_rows,
            issues=issues,
            stats=stats,
        )
        return self.result

    def export_csv(
        self,
        path: str | Path,
        *,
        delimiter: str = ",",
    ) -> None:
        result = self._require_result()
        target = Path(path)
        target.parent.mkdir(parents=True, exist_ok=True)

        with target.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=result.headers,
                delimiter=delimiter,
                extrasaction="ignore",
            )
            writer.writeheader()
            writer.writerows(result.rows)

    def export_json(self, path: str | Path) -> None:
        result = self._require_result()
        target = Path(path)
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(
            json.dumps(result.rows, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    def export_excel(self, path: str | Path) -> None:
        result = self._require_result()
        if Workbook is None:
            raise DataStudioError(self._t("engine_excel_dependency"))

        target = Path(path)
        target.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = self._t("summary_sheet")
        data_sheet = workbook.create_sheet(self._t("processed_sheet"))
        issues_sheet = workbook.create_sheet(self._t("issues_sheet"))

        self._write_summary_sheet(summary_sheet, result)
        self._write_data_sheet(data_sheet, result)
        self._write_issues_sheet(issues_sheet, result)

        try:
            workbook.save(target)
        except OSError as error:
            raise DataStudioError(
                self._t("engine_excel_save_error", error=error)
            )

    def export_issues(self, path: str | Path) -> None:
        result = self._require_result()
        target = Path(path)
        target.parent.mkdir(parents=True, exist_ok=True)

        with target.open("w", encoding="utf-8-sig", newline="") as handle:
            fieldnames = (
                self._t("issues_source_row"),
                self._t("issues_column"),
                self._t("issues_message"),
                self._t("issues_value"),
            )
            writer = csv.DictWriter(
                handle,
                fieldnames=fieldnames,
            )
            writer.writeheader()
            writer.writerows(
                {
                    fieldnames[0]: issue.source_row,
                    fieldnames[1]: issue.column,
                    fieldnames[2]: issue.message,
                    fieldnames[3]: issue.value,
                }
                for issue in result.issues
            )

    @staticmethod
    def save_profile(path: str | Path, config: ProcessingConfig) -> None:
        target = Path(path)
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(
            json.dumps(config.to_dict(), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    @staticmethod
    def load_profile(
        path: str | Path,
        *,
        language: str = "en",
    ) -> ProcessingConfig:
        source = Path(path)
        try:
            data = json.loads(source.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as error:
            raise DataStudioError(
                translate(language, "engine_profile_load_error", error=error)
            )

        if not isinstance(data, dict):
            raise DataStudioError(
                translate(language, "engine_profile_object_error")
            )
        return ProcessingConfig.from_dict(data)

    def _build_headers(
        self,
        config: ProcessingConfig,
    ) -> tuple[list[str], dict[str, str]]:
        transformed: list[str] = []

        for header in self.original_headers:
            renamed = config.column_renames.get(header, header).strip()
            value = normalize_header(renamed) if config.normalize_headers else renamed
            transformed.append(value or "column")

        headers = _make_unique(transformed)
        header_map = dict(zip(self.original_headers, headers))
        return headers, header_map

    @staticmethod
    def _clean_value(value: str, config: ProcessingConfig) -> str:
        return value.strip() if config.trim_whitespace else value

    @staticmethod
    def _matches_filter(value: str, mode: str, expected: str) -> bool:
        actual = value.casefold()
        wanted = expected.strip().casefold()

        if mode == "equals":
            return actual == wanted
        if mode == "not_equals":
            return actual != wanted
        if mode == "not_empty":
            return bool(value.strip())
        if mode == "empty":
            return not value.strip()
        return wanted in actual

    @staticmethod
    def _remove_duplicates(
        headers: list[str],
        rows: list[dict[str, str]],
    ) -> tuple[list[dict[str, str]], int]:
        seen: set[tuple[str, ...]] = set()
        unique_rows: list[dict[str, str]] = []

        for row in rows:
            signature = tuple(row.get(header, "") for header in headers)
            if signature in seen:
                continue
            seen.add(signature)
            unique_rows.append(row)

        return unique_rows, len(rows) - len(unique_rows)

    def _require_result(self) -> ProcessingResult:
        if self.result is None:
            raise DataStudioError(self._t("engine_apply_before_export"))
        return self.result

    def _write_summary_sheet(
        self,
        sheet,
        result: ProcessingResult,
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet["A1"] = self._t("summary_title")
        sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor="16324F")
        sheet.merge_cells("A1:D1")
        sheet["A1"].alignment = Alignment(vertical="center")
        sheet.row_dimensions[1].height = 32

        details = [
            (
                self._t("summary_source_file"),
                self.source_path.name if self.source_path else "",
            ),
            (self._t("summary_source_type"), self.source_kind.upper()),
            (
                self._t("summary_worksheet"),
                self.worksheet_name or self._t("summary_not_applicable"),
            ),
            (self._t("summary_source_rows"), result.stats["source_rows"]),
            (self._t("summary_output_rows"), result.stats["output_rows"]),
            (
                self._t("summary_empty_removed"),
                result.stats["empty_rows_removed"],
            ),
            (
                self._t("summary_filtered_removed"),
                result.stats["filtered_rows_removed"],
            ),
            (
                self._t("summary_duplicates_removed"),
                result.stats["duplicates_removed"],
            ),
            (
                self._t("summary_validation_issues"),
                result.stats["validation_issues"],
            ),
        ]
        for row_index, (label, value) in enumerate(details, start=3):
            sheet.cell(row=row_index, column=1, value=label)
            sheet.cell(row=row_index, column=2, value=value)

        label_range = sheet["A3:A11"]
        for (cell,) in label_range:
            cell.font = Font(bold=True, color="16324F")
            cell.fill = PatternFill("solid", fgColor="E8F0F7")

        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 34
        sheet.column_dimensions["C"].width = 4
        sheet.column_dimensions["D"].width = 4

    def _write_data_sheet(
        self,
        sheet,
        result: ProcessingResult,
    ) -> None:
        sheet.append(result.headers)
        for row in result.rows:
            sheet.append(
                [self._safe_excel_value(row.get(header, "")) for header in result.headers]
            )

        self._style_table_sheet(
            sheet,
            header_count=len(result.headers),
            row_count=max(1, len(result.rows) + 1),
            header_fill="16324F",
        )

    def _write_issues_sheet(
        self,
        sheet,
        result: ProcessingResult,
    ) -> None:
        headers = [
            self._t("issues_source_row"),
            self._t("issues_column"),
            self._t("issues_message"),
            self._t("issues_value"),
        ]
        sheet.append(headers)

        if result.issues:
            for issue in result.issues:
                sheet.append(
                    [
                        issue.source_row,
                        issue.column,
                        issue.message,
                        self._safe_excel_value(issue.value),
                    ]
                )
        else:
            sheet.append(["", "", self._t("issues_none"), ""])

        self._style_table_sheet(
            sheet,
            header_count=len(headers),
            row_count=max(2, len(result.issues) + 1),
            header_fill="B45309",
        )

    @staticmethod
    def _style_table_sheet(
        sheet,
        *,
        header_count: int,
        row_count: int,
        header_fill: str,
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = (
            f"A1:{get_column_letter(header_count)}{max(1, row_count)}"
        )

        border = Border(
            bottom=Side(style="thin", color="D9E2F3"),
        )
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=header_fill)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
        sheet.row_dimensions[1].height = 24

        for column_index in range(1, header_count + 1):
            values = [
                str(sheet.cell(row=row_index, column=column_index).value or "")
                for row_index in range(1, min(row_count, 200) + 1)
            ]
            width = min(max(max(map(len, values), default=0) + 2, 12), 42)
            sheet.column_dimensions[get_column_letter(column_index)].width = width

    @staticmethod
    def _cell_to_text(value) -> str:
        if value is None:
            return ""
        if isinstance(value, (datetime, date, time)):
            return value.isoformat()
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)

    @staticmethod
    def _safe_excel_value(value: str):
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value

    def _read_text(self, path: Path) -> tuple[str, str]:
        for encoding in SUPPORTED_ENCODINGS:
            try:
                return path.read_text(encoding=encoding), encoding
            except UnicodeDecodeError:
                continue
            except OSError as error:
                raise DataStudioError(
                    self._t("engine_text_read_error", error=error)
                )

        raise DataStudioError(self._t("engine_encoding_error"))

    @staticmethod
    def _detect_delimiter(text: str) -> str:
        sample = text[:8192]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=SUPPORTED_DELIMITERS)
            return dialect.delimiter
        except csv.Error:
            return ","
